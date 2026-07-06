#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B站排行榜爬虫 - Kivy Android App
数据源: bilibili.com/v/popular/rank/all
"""

import os
import sys
import time
import json
import threading
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.clock import Clock
from kivy.core.text import LabelBase
from kivy.utils import platform

if platform == "android":
    from android.permissions import request_permissions, Permission
    from plyer import sharing

# ============================================================
#  配置
# ============================================================
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.bilibili.com/v/popular/rank/all",
}

TIMEOUT = 20
API_RANKING = "https://api.bilibili.com/x/web-interface/ranking/v2"
API_HOTWORDS = "https://s.search.bilibili.com/main/hotword"


# ============================================================
#  爬虫逻辑（与 PC 版一致）
# ============================================================
def fetch_ranking():
    """获取B站全站排行榜"""
    try:
        resp = requests.get(API_RANKING, headers=HEADERS, timeout=TIMEOUT)
        if resp.status_code != 200:
            return None, f"HTTP {resp.status_code}"
        data = resp.json()
        if data.get("code") != 0:
            return None, f"API error: {data.get('message', 'unknown')}"
        return data.get("data", {}).get("list", []), None
    except Exception as e:
        return None, str(e)


def fetch_hotwords():
    """获取热搜词"""
    try:
        resp = requests.get(API_HOTWORDS, headers=HEADERS, timeout=TIMEOUT)
        if resp.status_code != 200:
            return []
        data = resp.json()
        if isinstance(data, dict):
            return data.get("list", [])
        return []
    except Exception:
        return []


def format_count(n):
    if n >= 100000000:
        return f"{n/100000000:.2f}亿"
    elif n >= 10000:
        return f"{n/10000:.1f}万"
    return str(n)


def format_duration(seconds):
    if seconds <= 0:
        return "-"
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"


def generate_excel(ranking, hotwords, filepath):
    """生成 Excel 文件"""
    wb = openpyxl.Workbook()

    hf = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    hfill = PatternFill(start_color="FB7299", end_color="FB7299", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tf = Font(name="微软雅黑", bold=True, size=14, color="333333")
    df = Font(name="微软雅黑", size=10)
    lf = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    r1 = Font(name="微软雅黑", bold=True, size=11, color="FB7299")
    r2 = Font(name="微软雅黑", bold=True, size=11, color="FF8C00")
    r3 = Font(name="微软雅黑", bold=True, size=11, color="FFC107")
    rn = Font(name="微软雅黑", size=10, color="666666")
    tb = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    ef = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

    ws = wb.active
    ws.title = "排行榜"
    headers = ["排名", "标题", "UP主", "播放量", "点赞", "弹幕", "评论", "时长", "综合评分", "链接"]
    widths = [8, 46, 16, 12, 10, 8, 8, 10, 12, 42]

    ws.merge_cells("A1:J1")
    tc = ws.cell(row=1, column=1, value=f"B站全站排行榜 ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    tc.font = tf
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    ws.row_dimensions[2].height = 28

    for ri, v in enumerate(ranking, 3):
        s = v.get("stat", {})
        row = [
            ri - 2, v.get("title", "").strip(), v.get("owner", {}).get("name", "未知"),
            format_count(s.get("view", 0)), format_count(s.get("like", 0)),
            format_count(s.get("danmaku", 0)), format_count(s.get("reply", 0)),
            format_duration(v.get("duration", 0)),
            f"{v.get('score', 0):.1f}" if v.get("score", 0) > 0 else "-",
            f"https://www.bilibili.com/video/{v.get('bvid', '')}",
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = tb
            if ci == 1:
                c.alignment = ha
                if ri - 2 == 1: c.font = r1
                elif ri - 2 == 2: c.font = r2
                elif ri - 2 == 3: c.font = r3
                else: c.font = rn
            elif ci == 2:
                c.font = df; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif ci == 10:
                c.font = lf; c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.font = df; c.alignment = ha
        if (ri - 3) % 2 == 1:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = ef

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"

    # Sheet 2: 热搜词
    if hotwords:
        ws2 = wb.create_sheet("热搜词")
        ws2.merge_cells("A1:D1")
        tc2 = ws2.cell(row=1, column=1, value=f"B站热搜词 ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
        tc2.font = tf; tc2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 36
        hw = ["排名", "热搜词", "热度值", "链接"]
        for ci, h in enumerate(hw, 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
        ws2.row_dimensions[2].height = 28
        for ri, w in enumerate(hotwords, 3):
            kw = w.get("keyword", "") or w.get("show_name", "")
            heat = w.get("heat_score", 0) or w.get("heat", 0) or 0
            url = f"https://search.bilibili.com/all?keyword={kw}"
            if not kw: continue
            row = [ri - 2, kw, format_count(heat), url]
            for ci, val in enumerate(row, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.border = tb
                if ci == 1:
                    c.alignment = ha
                    if ri - 2 == 1: c.font = r1
                    elif ri - 2 == 2: c.font = r2
                    elif ri - 2 == 3: c.font = r3
                    else: c.font = rn
                elif ci == 2:
                    c.font = df; c.alignment = Alignment(horizontal="left", vertical="center")
                elif ci == 4:
                    c.font = lf; c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.font = df; c.alignment = ha
            if (ri - 3) % 2 == 1:
                for ci in range(1, len(hw) + 1):
                    ws2.cell(row=ri, column=ci).fill = ef
        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 46
        ws2.column_dimensions["C"].width = 14
        ws2.column_dimensions["D"].width = 42
        ws2.freeze_panes = "A3"

    wb.save(filepath)
    return filepath


# ============================================================
#  Kivy App
# ============================================================
BILIBILI_PINK = (0.984, 0.447, 0.600, 1)  # #FB7299
DARK_BG = (0.12, 0.12, 0.14, 1)
CARD_BG = (0.18, 0.18, 0.22, 1)


class LogOutput(TextInput):
    """只读日志输出框"""
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.readonly = True
        self.background_color = DARK_BG
        self.foreground_color = (0.9, 0.9, 0.95, 1)
        self.font_size = 14
        self.padding = (12, 12)


class BilibiliRankingApp(App):
    """B站排行榜爬虫 App"""

    def build(self):
        self.title = "B站排行榜爬虫"
        self.icon = "icon.png"

        root = BoxLayout(orientation="vertical", spacing=10, padding=16)
        root.md_bg_color = DARK_BG

        # 标题
        title = Label(
            text="[b]B站排行榜爬虫[/b]",
            markup=True,
            font_size=22,
            color=BILIBILI_PINK,
            size_hint_y=0.08,
            halign="center",
        )
        root.add_widget(title)

        # 状态栏
        self.status_label = Label(
            text="就绪，点击下方按钮开始抓取",
            font_size=15,
            color=(0.7, 0.7, 0.8, 1),
            size_hint_y=0.06,
        )
        root.add_widget(self.status_label)

        # 日志区域
        scroll = ScrollView(size_hint_y=0.7)
        self.log = LogOutput()
        scroll.add_widget(self.log)
        root.add_widget(scroll)

        # 按钮行
        btn_layout = BoxLayout(
            orientation="horizontal",
            spacing=12,
            size_hint_y=0.1,
        )

        self.fetch_btn = Button(
            text="开始抓取",
            font_size=18,
            bold=True,
            background_color=BILIBILI_PINK,
            background_normal="",
            color=(1, 1, 1, 1),
        )
        self.fetch_btn.bind(on_press=self.start_fetch)

        self.share_btn = Button(
            text="分享文件",
            font_size=18,
            bold=True,
            background_color=(0.25, 0.25, 0.35, 1),
            background_normal="",
            color=(1, 1, 1, 1),
            disabled=True,
        )
        self.share_btn.bind(on_press=self.share_file)

        btn_layout.add_widget(self.fetch_btn)
        btn_layout.add_widget(self.share_btn)
        root.add_widget(btn_layout)

        self.excel_path = None
        self.ranking_data = []
        self.hotwords_data = []

        # Android 权限请求
        if platform == "android":
            try:
                request_permissions([
                    Permission.INTERNET,
                    Permission.READ_EXTERNAL_STORAGE,
                    Permission.WRITE_EXTERNAL_STORAGE,
                ])
            except Exception:
                pass

        return root

    def log_msg(self, msg):
        self.log.text += f"> {msg}\n"

    def set_status(self, text, is_error=False):
        color = (1, 0.3, 0.3, 1) if is_error else (0.7, 0.7, 0.8, 1)
        self.status_label.color = color
        self.status_label.text = text

    def start_fetch(self, instance):
        self.fetch_btn.disabled = True
        self.share_btn.disabled = True
        self.log.text = ""
        self.set_status("抓取中...")
        self.log_msg("=" * 40)
        self.log_msg(f"B站排行榜爬虫 v2.0")
        self.log_msg(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_msg("=" * 40)
        self.log_msg("")

        # 在后台线程抓取
        threading.Thread(target=self.do_fetch, daemon=True).start()

    def do_fetch(self):
        try:
            # 获取排行榜
            Clock.schedule_once(lambda dt: self.log_msg("[RANK] 正在获取排行榜数据..."))
            data, err = fetch_ranking()
            if err:
                Clock.schedule_once(lambda dt: self.log_msg(f"[X] 失败: {err}"))
                Clock.schedule_once(lambda dt: self.set_status("抓取失败", True))
                Clock.schedule_once(lambda dt: self._reset_btn())
                return

            self.ranking_data = data
            Clock.schedule_once(lambda dt: self.log_msg(f"[OK] 获取到 {len(data)} 条排行榜数据"))

            # 获取热搜词
            Clock.schedule_once(lambda dt: self.log_msg("[SEARCH] 正在获取热搜词..."))
            words = fetch_hotwords()
            self.hotwords_data = words
            Clock.schedule_once(lambda dt: self.log_msg(f"[OK] 获取到 {len(words)} 个热搜词"))

            # 显示 Top 5
            Clock.schedule_once(lambda dt: self.log_msg(""))
            Clock.schedule_once(lambda dt: self.log_msg("--- 排行榜 Top 5 ---"))
            for i, v in enumerate(data[:5], 1):
                title = v.get("title", "")[:30]
                play = format_count(v.get("stat", {}).get("view", 0))
                author = v.get("owner", {}).get("name", "?")
                Clock.schedule_once(
                    lambda dt, i=i, t=title, p=play, a=author:
                    self.log_msg(f"  #{i} {t}  {p}播放  UP:{a}")
                )

            # 生成 Excel
            Clock.schedule_once(lambda dt: self.log_msg(""))
            Clock.schedule_once(lambda dt: self.log_msg("[EXCEL] 正在生成 Excel 文件..."))

            # 保存路径
            if platform == "android":
                save_dir = os.path.join(os.environ.get("EXTERNAL_STORAGE", "/sdcard"), "Download")
            else:
                save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

            os.makedirs(save_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"B站排行榜_{timestamp}.xlsx"
            filepath = os.path.join(save_dir, filename)

            generate_excel(data, words, filepath)
            self.excel_path = filepath

            Clock.schedule_once(lambda dt: self.log_msg(f"[OK] 文件已保存!"))
            Clock.schedule_once(lambda dt: self.log_msg(f"      {filepath}"))

            # 完成
            Clock.schedule_once(lambda dt: self.set_status(f"完成! 已保存到 Download 目录"))
            Clock.schedule_once(lambda dt: self.log_msg(""))
            Clock.schedule_once(lambda dt: self.log_msg("=== 抓取完成 ==="))

        except Exception as e:
            Clock.schedule_once(lambda dt: self.log_msg(f"[X] 异常: {e}"))
            Clock.schedule_once(lambda dt: self.set_status("出错", True))

        Clock.schedule_once(lambda dt: self._reset_btn())

    def _reset_btn(self):
        self.fetch_btn.disabled = False
        if self.excel_path:
            self.share_btn.disabled = False

    def share_file(self, instance):
        if not self.excel_path or not os.path.exists(self.excel_path):
            self.set_status("没有可分享的文件", True)
            return

        try:
            if platform == "android":
                sharing.share(file_path=self.excel_path)
            else:
                # 桌面端直接打开目录
                os.startfile(os.path.dirname(self.excel_path))
        except Exception as e:
            self.set_status(f"分享失败: {e}", True)


if __name__ == "__main__":
    BilibiliRankingApp().run()
